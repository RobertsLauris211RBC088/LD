from openpyxl import Workbook, load_workbook 
wb=load_workbook("dip-225-1-practical-task-RobertsLauris211RBC088-main/tests/test1.xlsx")
ws=wb.active
max_row=ws.max_row
print(max_row)
for i in range(2,max_row+1):
    hours=ws['B'+str(i)].value
    rate=ws['C'+str(i)].value
    if(type(hours)!=str and type(rate)!=str):
        salary=float(hours)*float(rate)
        if salary>3000:
            ws['D'+str(i)].value=salary
            ws['E'+str(i)].value="malacis"
        else:
             ws['E'+str(i)].value="nožēlojami"
        print(salary)
wb.save('result.xlsx')
wb.close()